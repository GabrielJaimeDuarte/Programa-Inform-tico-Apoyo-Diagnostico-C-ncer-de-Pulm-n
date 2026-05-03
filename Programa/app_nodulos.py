import os
import re
import io
import cv2
import zipfile
import numpy as np
import streamlit as st
from PIL import Image
from ultralytics import YOLO
from collections import defaultdict
import tempfile
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

try:
    import pydicom
    PYDICOM_OK = True
except ImportError:
    PYDICOM_OK = False

st.set_page_config(page_title="Sistema de Deteccion de Nodulos Pulmonares",
    page_icon="🫁", layout="wide", initial_sidebar_state="expanded")

st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif;}
.stApp{background:#0D1117;}
.header-container{background:linear-gradient(135deg,#0f2027,#203a43,#2c5364);border-radius:16px;padding:2rem 2.5rem;margin-bottom:1.5rem;border:1px solid #1E3A5F;}
.header-title{color:white;font-size:1.9rem;font-weight:700;margin:0;letter-spacing:-0.02em;}
.header-sub{color:#7FBBDB;font-size:0.95rem;font-weight:300;margin-top:0.3rem;}
.header-badge{display:inline-block;background:rgba(0,180,255,0.15);border:1px solid rgba(0,180,255,0.3);color:#7FBBDB;font-size:0.78rem;padding:3px 10px;border-radius:20px;margin-right:6px;margin-top:8px;}
.metric-card{background:#161B22;border:1px solid #21262D;border-radius:12px;padding:1.2rem 1.5rem;text-align:center;}
.metric-value{font-size:2rem;font-weight:700;margin:0;line-height:1;}
.metric-label{color:#8B949E;font-size:0.8rem;margin-top:4px;text-transform:uppercase;letter-spacing:0.05em;}
.result-high{background:rgba(220,38,38,0.1);border:1px solid rgba(220,38,38,0.35);border-radius:12px;padding:1.2rem 1.5rem;margin:0.5rem 0;}
.result-medium{background:rgba(251,146,60,0.08);border:1px solid rgba(251,146,60,0.3);border-radius:12px;padding:1.2rem 1.5rem;margin:0.5rem 0;}
.result-low{background:rgba(251,191,36,0.07);border:1px solid rgba(251,191,36,0.25);border-radius:12px;padding:1.2rem 1.5rem;margin:0.5rem 0;}
.result-negative{background:rgba(34,197,94,0.08);border:1px solid rgba(34,197,94,0.25);border-radius:12px;padding:1.2rem 1.5rem;margin:0.5rem 0;}
.result-title{font-size:1.2rem;font-weight:700;margin:0;color:white;}
.result-sub{font-size:0.85rem;color:#8B949E;margin-top:3px;}
.nodule-badge{background:rgba(220,38,38,0.15);border:1px solid rgba(220,38,38,0.4);color:#F87171;font-size:0.75rem;font-weight:600;padding:2px 8px;border-radius:20px;}
.sospecha-badge{background:rgba(251,146,60,0.15);border:1px solid rgba(251,146,60,0.4);color:#FB923C;font-size:0.75rem;font-weight:600;padding:2px 8px;border-radius:20px;}
.hallazgo-badge{background:rgba(251,191,36,0.12);border:1px solid rgba(251,191,36,0.3);color:#FCD34D;font-size:0.75rem;font-weight:600;padding:2px 8px;border-radius:20px;}
.clean-badge{background:rgba(34,197,94,0.1);border:1px solid rgba(34,197,94,0.3);color:#4ADE80;font-size:0.75rem;font-weight:600;padding:2px 8px;border-radius:20px;}
.fp-badge{background:rgba(107,114,128,0.15);border:1px solid rgba(107,114,128,0.3);color:#9CA3AF;font-size:0.75rem;font-weight:600;padding:2px 8px;border-radius:20px;}
.nms-card-high{background:#0D1117;border:1px solid rgba(220,38,38,0.35);border-radius:10px;padding:0.8rem 1rem;margin:0.4rem 0;}
.nms-card-med{background:#0D1117;border:1px solid rgba(251,146,60,0.3);border-radius:10px;padding:0.8rem 1rem;margin:0.4rem 0;}
.nms-card-low{background:#0D1117;border:1px solid rgba(251,191,36,0.25);border-radius:10px;padding:0.8rem 1rem;margin:0.4rem 0;}
.fp-card{background:#0D1117;border:1px solid rgba(107,114,128,0.3);border-radius:10px;padding:0.8rem 1rem;margin:0.4rem 0;opacity:0.7;}
.pac-metric-card{background:#0D1117;border:1px solid #30363D;border-radius:10px;padding:1rem;text-align:center;}
.pac-metric-value{font-size:1.6rem;font-weight:700;margin:0;line-height:1;}
.pac-metric-label{color:#8B949E;font-size:0.75rem;margin-top:4px;}
.modo-ext{background:rgba(234,179,8,0.08);border:1px solid rgba(234,179,8,0.3);border-radius:10px;padding:0.8rem 1rem;margin:0.5rem 0;}
.stProgress>div>div{background-color:#3B82F6!important;}
div[data-testid="stSidebar"]{background:#0D1117;border-right:1px solid #21262D;}
.footer-note{text-align:center;color:#484F58;font-size:0.78rem;padding:1.5rem 0 0.5rem 0;border-top:1px solid #21262D;margin-top:2rem;}
</style>""", unsafe_allow_html=True)

RUTA_MODELO_A = r"D:/Modelo_Prueba_2/models/yolo_pulmon_nano/weights/best.pt"
RUTA_MODELO_B = r"D:/Modelo_Prueba_2/models/yolo_pulmon_nocturno2/weights/best.pt"
RUTA_MODELO_C = r"D:/Modelo_Prueba_2/models/yolo_pulmon_small_v22/weights/best.pt"

# Umbrales base para LUNA16 (dataset de entrenamiento)
CONF_NANO_BASE_LUNA    = 0.26; CONF_NANO_RUIDOSO_LUNA    = 0.40
CONF_FINE_BASE_LUNA    = 0.38; CONF_FINE_RUIDOSO_LUNA    = 0.52
CONF_SMALL_BASE_LUNA   = 0.35; CONF_SMALL_RUIDOSO_LUNA   = 0.50

# Umbrales ajustados para datasets externos (LIDC-IDRI, clinicos, etc.)
# Mas bajos porque el dominio es diferente y las confianzas son menores
CONF_NANO_BASE_EXT     = 0.18; CONF_NANO_RUIDOSO_EXT     = 0.30
CONF_FINE_BASE_EXT     = 0.28; CONF_FINE_RUIDOSO_EXT     = 0.40
CONF_SMALL_BASE_EXT    = 0.25; CONF_SMALL_RUIDOSO_EXT    = 0.38

if "resultados" not in st.session_state: st.session_state.resultados = None
if "analizado"  not in st.session_state: st.session_state.analizado  = False

@st.cache_resource
def cargar_modelos():
    with st.spinner("Cargando modelos..."):
        return YOLO(RUTA_MODELO_A), YOLO(RUTA_MODELO_B), YOLO(RUTA_MODELO_C)

# ==========================================
# VENTANA PULMONAR
# ==========================================
def ventana_pulmonar(imagen_hu, vmin=-1350, vmax=150):
    recortada   = np.clip(imagen_hu.astype(np.float32), vmin, vmax)
    normalizada = (recortada - vmin) / (vmax - vmin)
    return (normalizada * 255).astype(np.uint8)

def hu_a_slice_bgr(slice_2d_hu):
    gray = ventana_pulmonar(slice_2d_hu)
    return cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)

# ==========================================
# NPY → SLICES
# ==========================================
def npy_a_slices(archivo_npy, nombre_base):
    try:
        volumen = np.load(archivo_npy)
    except Exception as e:
        return None, f"Error al leer el NPY: {e}"
    if volumen.ndim == 2:
        volumen = volumen[np.newaxis, :, :]
    elif volumen.ndim != 3:
        return None, f"Formato inesperado: {volumen.ndim} dimensiones"
    slices = []
    for z in range(volumen.shape[0]):
        bgr = hu_a_slice_bgr(volumen[z, :, :])
        rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
        slices.append({"nombre": f"{nombre_base}_slice_{z:03d}.png",
                        "img_pil": Image.fromarray(rgb), "img_array": bgr})
    return slices, None

# ==========================================
# DICOM → SLICES
# ==========================================
def dicom_bytes_a_hu(dcm):
    pixel     = dcm.pixel_array.astype(np.float32)
    slope     = float(getattr(dcm, "RescaleSlope",     1))
    intercept = float(getattr(dcm, "RescaleIntercept", 0))
    return pixel * slope + intercept

def ordenar_dicoms(dcm_lista):
    def clave(item):
        dcm, nombre = item
        try:    return float(dcm.ImagePositionPatient[2])
        except: pass
        try:    return float(dcm.InstanceNumber)
        except: return nombre
    return sorted(dcm_lista, key=clave)

def zip_dicom_a_slices(zip_bytes, nombre_base):
    if not PYDICOM_OK:
        return None, "pydicom no instalado. Ejecuta: pip install pydicom"
    slices = []; errores = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        archivos = [n for n in zf.namelist()
                    if not n.endswith("/") and not os.path.basename(n).startswith(".")]
        dcm_lista = []
        for nombre in archivos:
            try:
                data = zf.read(nombre)
                dcm  = pydicom.dcmread(io.BytesIO(data), force=True)
                _    = dcm.pixel_array
                dcm_lista.append((dcm, nombre))
            except Exception as e:
                errores.append(str(e))
        if not dcm_lista:
            return None, f"No se encontraron DICOM válidos. {errores[:2]}"
        dcm_lista = ordenar_dicoms(dcm_lista)
        for z, (dcm, nombre) in enumerate(dcm_lista):
            try:
                hu  = dicom_bytes_a_hu(dcm)
                bgr = hu_a_slice_bgr(hu)
                rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
                slices.append({"nombre": f"{nombre_base}_slice_{z:03d}.png",
                               "img_pil": Image.fromarray(rgb), "img_array": bgr})
            except Exception as e:
                errores.append(f"slice {z}: {e}")
    return slices, None if slices else f"Sin slices convertidos. {errores[:2]}"

def dcm_sueltos_a_slices(archivos_dcm, nombre_base):
    if not PYDICOM_OK:
        return None, "pydicom no instalado."
    dcm_lista = []
    for arch in archivos_dcm:
        try:
            data = arch.read()
            dcm  = pydicom.dcmread(io.BytesIO(data), force=True)
            _    = dcm.pixel_array
            dcm_lista.append((dcm, arch.name))
        except: pass
    if not dcm_lista:
        return None, "No se pudieron leer los DICOM."
    dcm_lista = ordenar_dicoms(dcm_lista)
    slices = []
    for z, (dcm, nombre) in enumerate(dcm_lista):
        try:
            hu  = dicom_bytes_a_hu(dcm)
            bgr = hu_a_slice_bgr(hu)
            rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
            slices.append({"nombre": f"{nombre_base}_slice_{z:03d}.png",
                           "img_pil": Image.fromarray(rgb), "img_array": bgr})
        except: pass
    return slices, None if slices else "Sin slices convertidos."

# ==========================================
# DETECCIÓN
# ==========================================
def inferencia_tta(modelo, img_array, conf_umbral):
    cmax = 0.0; det = False; rimg = None
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tp = tmp.name
    cv2.imwrite(tp, img_array)
    for idx, v in enumerate([img_array, cv2.flip(img_array, 1), cv2.flip(img_array, 0)]):
        vp = tp.replace(".png", f"_v{idx}.png")
        cv2.imwrite(vp, v)
        r = modelo(vp, conf=conf_umbral, verbose=False)[0]
        if len(r.boxes) > 0:
            c = float(r.boxes.conf.max())
            if c > cmax: cmax = c; det = True
            if idx == 0 and det and rimg is None: rimg = r.plot()
        if os.path.exists(vp): os.remove(vp)
    if os.path.exists(tp): os.remove(tp)
    return det, rimg, cmax

def decidir_ensemble(da, db, dc, ca, cb, cc, ruidoso=False, modo_ext=False):
    # Para datasets externos bajamos los umbrales de decision del ensemble
    if modo_ext:
        if ca >= 0.42 or cb >= 0.48: base = da or db
        elif da and db: base = True
        elif da and ca >= 0.28: base = True
        elif db and cb >= 0.38: base = True
        else: base = False
        if not base: return False
        mab = max(ca, cb)
        if ruidoso:
            if not dc and cc < 0.01 and mab < 0.35: return False
        else:
            if not dc and cc < 0.02 and mab < 0.40: return False
        return True
    else:
        if ca >= 0.55 or cb >= 0.60: base = da or db
        elif da and db: base = True
        elif da and ca >= 0.38: base = True
        elif db and cb >= 0.52: base = True
        else: base = False
        if not base: return False
        mab = max(ca, cb)
        if ruidoso:
            if not dc and cc < 0.01 and mab < 0.45: return False
        else:
            if not dc and cc < 0.03 and mab < 0.52: return False
        return True

def analizar_slice(ma, mb, mc, img, cn, cf, cs, modo_ext=False):
    da, ia, ca = inferencia_tta(ma, img, cn)
    db, ib, cb = inferencia_tta(mb, img, cf)
    dc, ic, cc = inferencia_tta(mc, img, cs)
    hn = decidir_ensemble(da, db, dc, ca, cb, cc, modo_ext=modo_ext)
    cf2 = max(ca, cb, cc); ri = None
    for img2, c2 in [(ia, ca), (ib, cb), (ic, cc)]:
        if img2 is not None and c2 == cf2: ri = img2; break
    return hn, cf2, ri

# ==========================================
# CLASIFICACIÓN CLÍNICA
# ==========================================
def clasificar_hallazgo(n_slices, conf_max, modo_ext=False):
    cp = conf_max * 100
    # Para datos externos bajamos los umbrales de clasificacion
    umbral_alto = 55 if modo_ext else 60
    umbral_prob = 58 if modo_ext else 65
    umbral_med  = 45 if modo_ext else 55
    umbral_inc  = 50 if modo_ext else 59

    if n_slices >= 4:
        if cp >= umbral_alto:
            return ("alto","NODULO PROBABLE",
                    "Nodulo pulmonar probable — revision prioritaria por radiologo",
                    "#F87171","nms-card-high","nodule-badge")
        else:
            return ("medio","SOSPECHA MODERADA",
                    "Estructura sospechosa en multiples cortes — evaluacion recomendada",
                    "#FB923C","nms-card-med","sospecha-badge")
    elif n_slices in [2, 3]:
        if cp >= umbral_prob:
            return ("alto","NODULO PROBABLE",
                    "Nodulo probable en 2-3 cortes consecutivos con confianza alta — revision prioritaria",
                    "#F87171","nms-card-high","nodule-badge")
        elif cp >= umbral_med:
            return ("medio","SOSPECHA MODERADA",
                    "Estructura sospechosa en 2-3 cortes — evaluacion recomendada",
                    "#FB923C","nms-card-med","sospecha-badge")
        else:
            return ("descartado","PROBABLE ARTEFACTO",
                    "Estructura de baja confianza en pocos cortes — probable tejido vascular o artefacto",
                    "#9CA3AF","fp-card","fp-badge")
    else:
        if cp >= umbral_inc:
            return ("bajo","HALLAZGO INCIDENTAL",
                    "Hallazgo en corte unico con confianza aceptable — verificacion manual recomendada",
                    "#FCD34D","nms-card-low","hallazgo-badge")
        else:
            return ("descartado","PROBABLE ARTEFACTO",
                    "Deteccion en corte unico de baja confianza — probable tejido vascular o artefacto",
                    "#9CA3AF","fp-card","fp-badge")

def nms_3d_con_validacion(resultados, gap_max=3, conf_1slice=0.59,
                           conf_2_3slices=0.55, modo_ext=False):
    positivos = [(i, r["confianza"], r["nombre"])
                 for i, r in enumerate(resultados) if r["hay_nodulo"]]
    if not positivos: return [], []

    grupos = []; ga = [positivos[0]]
    for j in range(1, len(positivos)):
        if positivos[j][0] - positivos[j-1][0] - 1 <= gap_max:
            ga.append(positivos[j])
        else:
            grupos.append(ga); ga = [positivos[j]]
    grupos.append(ga)

    validos = []; descartados = []
    for i, g in enumerate(grupos):
        idx   = [x[0] for x in g]; confs = [x[1] for x in g]; noms = [x[2] for x in g]
        n_det = len(idx); nr = idx[-1] - idx[0] + 1
        cm    = max(confs); cp = float(np.mean(confs))
        im    = confs.index(cm); dm = round(nr * 1.25, 1)
        if dm < 6:   ts = "Pequeno (<6mm)"
        elif dm < 10: ts = "Mediano (6-10mm)"
        elif dm < 20: ts = "Grande (10-20mm)"
        else:         ts = "Muy grande (>20mm)"
        nivel, etiq, desc_larga, color, css, badge = clasificar_hallazgo(n_det, cm, modo_ext)
        nod = {"id": i+1, "slices": idx, "nombres": noms,
               "slice_inicio": idx[0], "slice_fin": idx[-1],
               "slice_central": idx[im], "nombre_central": noms[im],
               "conf_max": cm, "conf_prom": cp, "n_slices": n_det,
               "n_slices_rango": nr, "diametro_est_mm": dm, "tamano_estimado": ts,
               "nivel": nivel, "etiqueta": etiq, "descripcion": desc_larga,
               "color": color, "css": css, "badge": badge}
        if nivel == "descartado":
            nod["razon_descarte"] = desc_larga; descartados.append(nod)
        else:
            nod["razon_validacion"] = desc_larga; validos.append(nod)

    # Validacion secundaria con umbrales del sidebar
    c1_eff  = conf_1slice  * (0.85 if modo_ext else 1.0)
    c23_eff = conf_2_3slices * (0.85 if modo_ext else 1.0)
    validos_final = []; descartados_extra = []
    for nod in validos:
        n = nod["n_slices"]; cm = nod["conf_max"]
        if n == 1 and cm < c1_eff:
            nod["nivel"] = "descartado"
            nod["razon_descarte"] = f"1 slice, conf {cm*100:.1f}% < {c1_eff*100:.0f}% requerido"
            descartados_extra.append(nod)
        elif n <= 3 and cm < c23_eff:
            nod["nivel"] = "descartado"
            nod["razon_descarte"] = f"{n} slices, conf {cm*100:.1f}% < {c23_eff*100:.0f}% requerido"
            descartados_extra.append(nod)
        else:
            validos_final.append(nod)

    descartados_total = descartados + descartados_extra
    for j, nod in enumerate(validos_final, 1): nod["id"] = j
    return validos_final, descartados_total

def calcular_metricas_pac(resultados):
    pacs = defaultdict(list)
    for r in resultados:
        p   = r["nombre"].replace(".png","").replace(".jpg","").split("_")
        pid = f"paciente_{p[1]}" if len(p) >= 2 and p[0] == "paciente" else "paciente_anonimo"
        pacs[pid].append(r)
    m = {}
    for pid, sl in pacs.items():
        pos = [s for s in sl if s["hay_nodulo"]]; n = len(pos)
        c   = [s["confianza"] for s in pos]
        m[pid] = {"n_total": len(sl), "n_positivos": n, "n_limpios": len(sl)-n,
                  "conf_max": max(c) if c else 0.0, "conf_prom": float(np.mean(c)) if c else 0.0,
                  "tiene_nodulo": n > 0, "pct_afectado": n/len(sl)*100 if sl else 0}
    return m

def generar_excel(resultados, mpac, validos, descartados):
    wb  = openpyxl.Workbook()
    hf  = PatternFill("solid", fgColor="0F3460")
    hft = Font(bold=True, color="FFFFFF", size=10)
    borde = Border(left=Side(style="thin",color="333333"),right=Side(style="thin",color="333333"),
                   top=Side(style="thin",color="333333"),bottom=Side(style="thin",color="333333"))
    def hdr(ws, row, cols):
        for col, v in enumerate(cols, 1):
            c = ws.cell(row=row, column=col, value=v); c.font=hft; c.fill=hf; c.border=borde
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20

    ws0 = wb.active; ws0.title = "Hallazgos validos"; ws0.sheet_view.showGridLines = False
    ws0.merge_cells("A1:I1")
    ws0["A1"] = f"Hallazgos Validos — NMS 3D ({len(validos)} hallazgo(s))"
    ws0["A1"].font = Font(bold=True,color="FFFFFF",size=13); ws0["A1"].fill = hf
    ws0["A1"].alignment = Alignment(horizontal="center",vertical="center"); ws0.row_dimensions[1].height = 28
    hdr(ws0, 3, ["N","Clasificacion clinica","Rango slices","Slice central",
                 "Conf max (%)","Diam est mm","Tamano","N slices","Descripcion"])
    for i, n in enumerate(validos, 4):
        fd = PatternFill("solid", fgColor="1A0808" if i%2==0 else "2A0A0A")
        cc = ("F87171" if n["nivel"]=="alto" else "FB923C" if n["nivel"]=="medio" else "FCD34D")
        for col, val in enumerate([n['id'],n['etiqueta'],
            f"slice_{n['slice_inicio']:03d} a slice_{n['slice_fin']:03d}",
            f"slice_{n['slice_central']:03d}",
            round(n['conf_max']*100,1),n['diametro_est_mm'],
            n['tamano_estimado'],n['n_slices'],n['descripcion']], 1):
            c = ws0.cell(row=i,column=col,value=val); c.fill=fd; c.border=borde
            c.alignment = Alignment(horizontal="left" if col in [2,3,9] else "center",
                                    vertical="center", wrap_text=(col==9))
            c.font = Font(color=cc if col==2 else "E6EDF3", size=10, bold=(col==2))
        ws0.row_dimensions[i].height = 18
    for col, a in enumerate([6,22,28,20,16,14,18,10,45], 1):
        ws0.column_dimensions[get_column_letter(col)].width = a

    ws_fp = wb.create_sheet("Descartados"); ws_fp.sheet_view.showGridLines = False
    ws_fp.merge_cells("A1:G1")
    ws_fp["A1"] = f"Descartados — probables artefactos ({len(descartados)})"
    ws_fp["A1"].font = Font(bold=True,color="FFFFFF",size=13); ws_fp["A1"].fill = hf
    ws_fp["A1"].alignment = Alignment(horizontal="center"); ws_fp.row_dimensions[1].height = 28
    hdr(ws_fp, 3, ["Grupo","Rango slices","Slice central","Conf max (%)","N slices","Tamano","Razon descarte"])
    for i, n in enumerate(descartados, 4):
        fd = PatternFill("solid", fgColor="121212")
        for col, val in enumerate([f"Grupo {i-3}",
            f"slice_{n['slice_inicio']:03d} a slice_{n['slice_fin']:03d}",
            f"slice_{n['slice_central']:03d}",
            round(n['conf_max']*100,1),n['n_slices'],
            n['tamano_estimado'],n.get('razon_descarte','')], 1):
            c = ws_fp.cell(row=i,column=col,value=val); c.fill=fd; c.border=borde
            c.alignment = Alignment(horizontal="left" if col in [2,7] else "center",vertical="center")
            c.font = Font(color="888888",size=10)
        ws_fp.row_dimensions[i].height = 16
    for col, a in enumerate([10,28,20,14,10,16,45], 1):
        ws_fp.column_dimensions[get_column_letter(col)].width = a

    ws1 = wb.create_sheet("Detalle slices"); ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A1:F1")
    ws1["A1"] = f"Reporte por Slice — {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1["A1"].font = Font(bold=True,color="FFFFFF",size=13); ws1["A1"].fill = hf
    ws1["A1"].alignment = Alignment(horizontal="center"); ws1.row_dimensions[1].height = 26
    hdr(ws1, 3, ["#","Archivo","Clasificacion","Confianza (%)","Nivel","Interpretacion"])
    for i, r in enumerate(resultados, 1):
        row = i+3; cp = r["confianza"]*100
        if r["hay_nodulo"]:
            if cp>=60: cl="Estructura sospechosa"; nv="Alta"; ob="Revisar con radiologo"
            elif cp>=45: cl="Hallazgo incidental"; nv="Media"; ob="Correlacion clinica recomendada"
            else: cl="Posible artefacto"; nv="Baja"; ob="Probable tejido vascular"
        else:
            cl="Sin hallazgos"; nv="-"; ob="Corte limpio"
        fd = PatternFill("solid", fgColor="1A0A0A" if r["hay_nodulo"] else "0A1A0A")
        for col, val in enumerate([i,r["nombre"],cl,round(cp,2),nv,ob], 1):
            c = ws1.cell(row=row,column=col,value=val); c.fill=fd; c.border=borde
            c.alignment = Alignment(horizontal="left" if col in [2,3,6] else "center",vertical="center")
            if col==3 and r["hay_nodulo"]:
                c.font = Font(bold=True,color="F87171" if cp>=60 else "FB923C" if cp>=45 else "9CA3AF")
            else: c.font = Font(color="E6EDF3",size=10)
        ws1.row_dimensions[row].height = 16
    for col, a in enumerate([5,35,22,14,10,35], 1):
        ws1.column_dimensions[get_column_letter(col)].width = a

    ws3 = wb.create_sheet("Metricas modelo"); ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:C1"); ws3["A1"] = "Metricas V31 Triple Ensemble"
    ws3["A1"].font = Font(bold=True,color="FFFFFF",size=13); ws3["A1"].fill = hf
    ws3["A1"].alignment = Alignment(horizontal="center"); ws3.row_dimensions[1].height = 26
    hdr(ws3, 3, ["Metrica","Por nodulo","Por paciente"])
    for i, (m,n,p) in enumerate([("Exactitud","93.20%","93.26%"),("Sensibilidad","88.76%","94.59%"),
        ("Precision","86.09%","94.59%"),("Especificidad","94.80%","91.04%"),("F1-Score","87.40%","94.59%")], 4):
        f = PatternFill("solid", fgColor="161B22" if i%2==0 else "0D1117")
        for col, val in enumerate([m,n,p], 1):
            c = ws3.cell(row=i,column=col,value=val); c.fill=f; c.border=borde
            c.font = Font(color="E6EDF3",size=11,bold=(col==1))
            c.alignment = Alignment(horizontal="left" if col==1 else "center",vertical="center")
        ws3.row_dimensions[i].height = 20
    ws3.column_dimensions["A"].width=18; ws3.column_dimensions["B"].width=16; ws3.column_dimensions["C"].width=16
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

def generar_txt(resultados, mpac, validos, descartados, modo_ext=False):
    ahora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    L = ["="*60,"  REPORTE — ANALISIS PULMONAR",f"  Fecha: {ahora}",
         f"  Sistema: YOLOv8 Triple Ensemble V31 + NMS 3D",
         f"  Modo: {'Dataset externo' if modo_ext else 'LUNA16'}",
         "="*60,""]
    L += [f"HALLAZGOS VALIDOS: {len(validos)}","-"*40]
    if validos:
        for n in validos:
            L += [f"  [{n['etiqueta']}] Hallazgo {n['id']}:",
                  f"    Rango     : slice_{n['slice_inicio']:03d} -> slice_{n['slice_fin']:03d}",
                  f"    Central   : slice_{n['slice_central']:03d}",
                  f"    Confianza : {n['conf_max']*100:.1f}%",
                  f"    Diametro  : ~{n['diametro_est_mm']}mm ({n['tamano_estimado']})",
                  f"    Descripcion: {n['descripcion']}",""]
    else:
        L.append("  Sin hallazgos significativos.")
    L += ["","DESCARTADOS: "+str(len(descartados)),"-"*40]
    if descartados:
        for n in descartados:
            L += [f"  slice_{n['slice_inicio']:03d}: {n.get('razon_descarte','')}"]
    L += ["","="*60,"  METRICAS DEL MODELO","="*60,
          "  Exactitud    93.20% / 93.26%","  Sensibilidad 88.76% / 94.59%",
          "  Precision   86.09% / 94.59%","  Especific.  94.80% / 91.04%",
          "  F1-Score    87.40% / 94.59%","",
          "  NOTA: Herramienta de apoyo al diagnostico.",
          "  Confirmar hallazgos con radiologo certificado."]

    # ── NUEVO: lista de archivos analizados para el comparador ──
    L += ["","="*60,"  ARCHIVOS_ANALIZADOS","="*60]
    for r in resultados:
        hay = "1" if r["hay_nodulo"] else "0"
        conf = f"{r['confianza']*100:.1f}"
        L.append(f"  {r['nombre']}|{hay}|{conf}")
    L.append("="*60)

    return "\n".join(L)

# ==========================================
# HEADER
# ==========================================
st.markdown("""<div class="header-container">
<p class="header-title">🫁 Sistema de Analisis Pulmonar</p>
<p class="header-sub">Proyecto de Grado — YOLOv8 Triple Ensemble + NMS 3D con clasificacion clinica</p>
<span class="header-badge">Triple Ensemble</span><span class="header-badge">TTA</span>
<span class="header-badge">NMS 3D</span><span class="header-badge">PNG · NPY · DICOM</span>
<span class="header-badge">LUNA16 · LIDC-IDRI</span></div>""", unsafe_allow_html=True)

# ==========================================
# SIDEBAR
# ==========================================
with st.sidebar:
    st.markdown("### ⚙️ Configuracion"); st.markdown("---")

    # MODO DE DATASET — nuevo
    st.markdown("### 🌐 Modo de dataset")
    modo_ext = st.toggle("Dataset externo (LIDC-IDRI / Clinico)",
        value=False,
        help="Activa umbrales reducidos para datasets diferentes a LUNA16. "
             "Usar cuando se analizan datos de otros equipos CT o instituciones.")

    if modo_ext:
        st.markdown("""<div class="modo-ext">
        <b style="color:#EAB308;">⚡ Modo dataset externo activo</b><br>
        <span style="color:#FDE68A;font-size:0.82rem;">
        Umbrales reducidos para compensar diferencias de dominio.<br>
        Puede generar mas detecciones — revisar con criterio clinico.
        </span></div>""", unsafe_allow_html=True)
    else:
        st.markdown("""<div style="background:rgba(59,130,246,0.06);border:1px solid rgba(59,130,246,0.2);
        border-radius:8px;padding:0.6rem;color:#93C5FD;font-size:0.82rem;">
        ✅ Modo LUNA16 — umbrales optimizados para el dataset de entrenamiento
        </div>""", unsafe_allow_html=True)

    st.markdown("---")
    umbral = st.slider("Umbral confianza minima", 0.10, 0.60,
                   0.18 if modo_ext else 0.26, 0.01,
                   help="Nivel mínimo de certeza para que el modelo reporte una detección. "
                        "Bajar = más sensible pero más falsas alarmas. "
                        "Subir = menos falsas alarmas pero puede perder nódulos.")
    st.markdown("---")
    st.markdown("### 🧠 NMS 3D — Agrupador")
    gap = st.slider("Gap maximo entre slices", 1, 8, 3, 1,
                help="Cuántos slices limpios se toleran dentro de un mismo hallazgo. "
                     "Gap=3 significa que si hay 3 slices sin detección entre dos detecciones, "
                     "se consideran parte del mismo nódulo.")
    st.markdown("""<div style="background:rgba(59,130,246,0.08);border:1px solid rgba(59,130,246,0.2);
    border-radius:8px;padding:0.6rem;margin-top:0.5rem;color:#93C5FD;font-size:0.8rem;">
    <b>Gap=1:</b> estricto &nbsp; <b>Gap=3:</b> recomendado &nbsp; <b>Gap=6+:</b> permisivo
    </div>""", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### 🏥 Criterios de validacion")
    conf_1_def  = 0.50 if modo_ext else 0.59
    conf_23_def = 0.45 if modo_ext else 0.55
    conf_1 = st.slider("Conf. minima — 1 corte solo", 0.30, 0.80, conf_1_def, 0.01,
                   help="Umbral especial para detecciones en un único slice. "
                        "Como un slice suelto suele ser tejido vascular, se exige más confianza.")
    conf_23 = st.slider("Conf. minima — 2 o 3 cortes", 0.25, 0.75, conf_23_def, 0.01,
                    help="Umbral para hallazgos en 2-3 slices consecutivos. "
                         "Menos restrictivo que 1 corte porque múltiples cortes "
                         "son más indicativos de una estructura real.")
    st.markdown("""<div style="background:rgba(34,197,94,0.06);border:1px solid rgba(34,197,94,0.2);
    border-radius:8px;padding:0.6rem;margin-top:0.5rem;color:#86EFAC;font-size:0.8rem;">
    <b>4+ cortes:</b> siempre reportado
    </div>""", unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### 📋 Leyenda")
    st.markdown("🔴 **Nodulo probable** — revision prioritaria\n"
                "🟠 **Sospecha moderada** — evaluar\n"
                "🟡 **Hallazgo incidental** — no concluyente\n"
                "⚪ **Probable artefacto** — tejido vascular\n"
                "✅ **Sin hallazgos** — corte limpio")
    st.markdown("---")
    st.markdown("### 📊 Metricas (LUNA16)")
    st.markdown("|Metrica|Nodulo|Paciente|\n|---|---|---|\n"
                "|Exactitud|93.20%|93.26%|\n|Sensibilidad|88.76%|94.59%|\n"
                "|Precision|86.09%|94.59%|\n|Especificidad|94.80%|91.04%|\n"
                "|F1-Score|87.40%|94.59%|")

# ==========================================
# CARGA — 3 TABS
# ==========================================
st.markdown("### 📂 Cargar archivos del paciente")
tab_png, tab_npy, tab_dcm = st.tabs([
    "📸 PNG / JPG",
    "🧬 NPY (volumen)",
    "🏥 DICOM (ZIP o archivos .dcm)"
])
archivos_procesados = []

with tab_png:
    st.markdown('<div style="color:#8B949E;font-size:0.85rem;margin-bottom:0.8rem;">'
                'Slices PNG exportados. Ctrl+Click para seleccionar varios.</div>',
                unsafe_allow_html=True)
    uk = f"uploader_{st.session_state.get('upload_key',0)}"
    archivos_png = st.file_uploader("Selecciona los cortes PNG",
        type=["png","jpg","jpeg"], accept_multiple_files=True, key=uk)
    if archivos_png:
        for arch in sorted(archivos_png, key=lambda x: x.name):
            img_pil   = Image.open(arch).convert("RGB")
            img_array = cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)
            archivos_procesados.append({"nombre":arch.name,"img_pil":img_pil,"img_array":img_array})

with tab_npy:
    st.markdown("""<div style="background:rgba(99,102,241,0.08);border:1px solid rgba(99,102,241,0.25);
    border-radius:10px;padding:0.8rem 1rem;margin-bottom:0.8rem;color:#A5B4FC;font-size:0.85rem;">
    🧬 Volumen 3D .npy — ventana pulmonar (-1350/+150 HU) aplicada automaticamente.
    </div>""", unsafe_allow_html=True)
    archivo_npy = st.file_uploader("Selecciona el archivo .npy",
        type=["npy"], accept_multiple_files=False, key="npy_uploader")
    if archivo_npy:
        nombre_base = archivo_npy.name.replace(".npy","")
        with st.spinner(f"Convirtiendo {archivo_npy.name}..."):
            with tempfile.NamedTemporaryFile(suffix=".npy", delete=False) as tmp_npy:
                tmp_npy.write(archivo_npy.read()); tmp_path = tmp_npy.name
            slices_npy, error = npy_a_slices(tmp_path, nombre_base)
            os.remove(tmp_path)
        if error: st.error(error)
        else:
            st.markdown(f'<div style="background:rgba(34,197,94,0.08);border:1px solid '
                        f'rgba(34,197,94,0.25);border-radius:10px;padding:0.8rem 1rem;'
                        f'color:#86EFAC;font-size:0.85rem;">✅ <b>{archivo_npy.name}</b> '
                        f'— <b>{len(slices_npy)} slices</b> listos.</div>',
                        unsafe_allow_html=True)
            archivos_procesados.extend(slices_npy)

with tab_dcm:
    if not PYDICOM_OK:
        st.warning("pydicom no instalado. Ejecuta: pip install pydicom")
    else:
        st.markdown("""<div style="background:rgba(20,184,166,0.08);border:1px solid rgba(20,184,166,0.25);
        border-radius:10px;padding:0.8rem 1rem;margin-bottom:0.8rem;color:#5EEAD4;font-size:0.85rem;">
        🏥 <b>Compatible con cualquier dataset DICOM estandar.</b><br>
        Slices ordenados por posicion anatomica real (ImagePositionPatient).<br>
        Funciona con LUNA16, LIDC-IDRI, NLST y datos clinicos.<br><br>
        <b>Opcion A:</b> ZIP con todos los .dcm del paciente (recomendado)<br>
        <b>Opcion B:</b> Archivos .dcm directos (Ctrl+Click para varios)
        </div>""", unsafe_allow_html=True)
        dcm_modo   = st.radio("Modo DICOM", ["ZIP (recomendado)","Archivos .dcm directos"], horizontal=True)
        nombre_dcm = st.text_input("Nombre del paciente", value="paciente", key="dcm_nombre")
        if dcm_modo == "ZIP (recomendado)":
            archivo_zip = st.file_uploader("Sube el ZIP con los DICOM",
                type=["zip"], accept_multiple_files=False, key="zip_uploader")
            if archivo_zip:
                with st.spinner(f"Procesando {archivo_zip.name}..."):
                    slices_dcm, error = zip_dicom_a_slices(archivo_zip.read(), nombre_dcm)
                if error: st.error(error)
                elif slices_dcm:
                    st.markdown(f'<div style="background:rgba(34,197,94,0.08);border:1px solid '
                                f'rgba(34,197,94,0.25);border-radius:10px;padding:0.8rem 1rem;'
                                f'color:#86EFAC;font-size:0.85rem;">✅ <b>{len(slices_dcm)} slices</b> '
                                f'ordenados. Listos.</div>', unsafe_allow_html=True)
                    archivos_procesados.extend(slices_dcm)
        else:
            archivos_dcm = st.file_uploader("Selecciona los .dcm",
                type=["dcm","dicom"], accept_multiple_files=True, key="dcm_uploader")
            if archivos_dcm:
                with st.spinner(f"Procesando {len(archivos_dcm)} archivos..."):
                    slices_dcm, error = dcm_sueltos_a_slices(archivos_dcm, nombre_dcm)
                if error: st.error(error)
                elif slices_dcm:
                    st.markdown(f'<div style="background:rgba(34,197,94,0.08);border:1px solid '
                                f'rgba(34,197,94,0.25);border-radius:10px;padding:0.8rem 1rem;'
                                f'color:#86EFAC;font-size:0.85rem;">✅ <b>{len(slices_dcm)} slices</b> '
                                f'listos.</div>', unsafe_allow_html=True)
                    archivos_procesados.extend(slices_dcm)
        st.markdown('<div style="color:#484F58;font-size:0.78rem;margin-top:0.5rem;">'
                    '💡 Windows: selecciona los .dcm → clic derecho → Comprimir en ZIP</div>',
                    unsafe_allow_html=True)

# ==========================================
# ANÁLISIS
# ==========================================
if archivos_procesados:
    nt = len(archivos_procesados)
    modo_txt = "Dataset externo" if modo_ext else "LUNA16"
    st.markdown(f'<div style="background:#161B22;border:1px solid #21262D;border-radius:10px;'
                f'padding:0.8rem 1.2rem;margin:0.5rem 0;color:#E6EDF3;font-size:0.9rem;">'
                f'📁 <b>{nt} cortes listos</b> — Modo: <b style="color:{"#EAB308" if modo_ext else "#3B82F6"};">'
                f'{modo_txt}</b></div>', unsafe_allow_html=True)

    c1, c2, c3 = st.columns([3,1.2,1.2])
    with c1: ba = st.button("🔬 Analizar paciente completo", use_container_width=True)
    with c2: bl = st.button("🗑️ Limpiar resultados",        use_container_width=True)
    with c3: bi = st.button("❌ Borrar archivos",            use_container_width=True)

    if bi:
        st.session_state.resultados=None; st.session_state.analizado=False
        st.session_state["upload_key"]=st.session_state.get("upload_key",0)+1; st.rerun()
    if bl:
        st.session_state.resultados=None; st.session_state.analizado=False; st.rerun()

    if ba:
        ma, mb, mc = cargar_modelos()
        # Seleccionar umbrales segun modo
        if modo_ext:
            cn2 = CONF_NANO_BASE_EXT  + (umbral - 0.18)
            cf2 = CONF_FINE_BASE_EXT  + (umbral - 0.18)
            cs2 = CONF_SMALL_BASE_EXT + (umbral - 0.18)
        else:
            cn2 = umbral; cf2 = umbral + 0.12; cs2 = umbral + 0.09

        res = []; bar = st.progress(0, text="Analizando cortes...")
        for idx, item in enumerate(archivos_procesados):
            hn, cf_, ir = analizar_slice(ma, mb, mc, item["img_array"], cn2, cf2, cs2, modo_ext)
            res.append({"nombre": item["nombre"], "hay_nodulo": hn, "confianza": cf_,
                        "img_orig": item["img_pil"],
                        "img_result": Image.fromarray(cv2.cvtColor(ir, cv2.COLOR_BGR2RGB))
                                      if ir is not None else item["img_pil"]})
            bar.progress((idx+1)/nt, text=f"Analizando {idx+1}/{nt}: {item['nombre']}")
        bar.empty(); st.session_state.resultados=res; st.session_state.analizado=True

    if st.session_state.analizado and st.session_state.resultados:
        res = st.session_state.resultados
        sp  = [r for r in res if r["hay_nodulo"]]
        sl_ = [r for r in res if not r["hay_nodulo"]]
        np_ = len(sp); cp_ = [r["confianza"] for r in sp]
        cmg = max(cp_) if cp_ else 0
        ntr = len(res); mpac = calcular_metricas_pac(res)
        validos, descartados = nms_3d_con_validacion(res, gap_max=gap,
            conf_1slice=conf_1, conf_2_3slices=conf_23, modo_ext=modo_ext)
        niveles   = {"alto":3,"medio":2,"bajo":1,"descartado":0}
        nivel_max = max((niveles.get(n["nivel"],0) for n in validos), default=0)

        if modo_ext:
            st.markdown("""<div class="modo-ext">
            ⚡ <b style="color:#EAB308;">Analisis en modo dataset externo</b>
            <span style="color:#FDE68A;font-size:0.82rem;"> — umbrales reducidos para compensar diferencias de dominio respecto a LUNA16.</span>
            </div>""", unsafe_allow_html=True)

        st.markdown("---"); st.markdown("## 📋 Resultado del Analisis")
        if nivel_max == 3:
            st.markdown(f'<div class="result-high"><p class="result-title">🔴 NODULO(S) PROBABLE(S)</p>'
                        f'<p class="result-sub"><b>{len(validos)} hallazgo(s)</b> — revision prioritaria. '
                        f'Conf max: <b>{cmg*100:.1f}%</b></p></div>', unsafe_allow_html=True)
        elif nivel_max == 2:
            st.markdown(f'<div class="result-medium"><p class="result-title">🟠 SOSPECHA MODERADA</p>'
                        f'<p class="result-sub"><b>{len(validos)} estructura(s)</b> — evaluacion recomendada.</p></div>',
                        unsafe_allow_html=True)
        elif nivel_max == 1:
            st.markdown(f'<div class="result-low"><p class="result-title">🟡 HALLAZGO INCIDENTAL</p>'
                        f'<p class="result-sub"><b>{len(validos)} hallazgo(s)</b> — no concluyente.</p></div>',
                        unsafe_allow_html=True)
        elif np_ > 0 and len(validos) == 0:
            st.markdown('<div style="background:rgba(107,114,128,0.08);border:1px solid rgba(107,114,128,0.25);'
                        'border-radius:12px;padding:1.2rem 1.5rem;margin:0.5rem 0;">'
                        '<p class="result-title" style="color:#9CA3AF;">⚪ DETECCIONES DESCARTADAS</p>'
                        '<p class="result-sub">Ninguna supero los criterios de validacion.</p></div>',
                        unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="result-negative"><p class="result-title">✅ SIN HALLAZGOS SIGNIFICATIVOS</p>'
                        f'<p class="result-sub">Los {ntr} cortes no muestran estructuras sospechosas.</p></div>',
                        unsafe_allow_html=True)

        st.markdown("#### Resumen")
        cols5 = st.columns(5)
        n_alto  = sum(1 for n in validos if n["nivel"]=="alto")
        n_medio = sum(1 for n in validos if n["nivel"]=="medio")
        n_bajo  = sum(1 for n in validos if n["nivel"]=="bajo")
        with cols5[0]: st.markdown(f'<div class="metric-card"><p class="metric-value" style="color:#F87171;">{n_alto}</p><p class="metric-label">Nodulos probables</p></div>',unsafe_allow_html=True)
        with cols5[1]: st.markdown(f'<div class="metric-card"><p class="metric-value" style="color:#FB923C;">{n_medio}</p><p class="metric-label">Sospechas</p></div>',unsafe_allow_html=True)
        with cols5[2]: st.markdown(f'<div class="metric-card"><p class="metric-value" style="color:#FCD34D;">{n_bajo}</p><p class="metric-label">Hallazgos incid.</p></div>',unsafe_allow_html=True)
        with cols5[3]: st.markdown(f'<div class="metric-card"><p class="metric-value" style="color:#9CA3AF;">{len(descartados)}</p><p class="metric-label">Descartados</p></div>',unsafe_allow_html=True)
        with cols5[4]: st.markdown(f'<div class="metric-card"><p class="metric-value" style="color:#3B82F6;">{cmg*100:.1f}%</p><p class="metric-label">Conf maxima</p></div>',unsafe_allow_html=True)

        if validos:
            st.markdown("---"); st.markdown("### 🔬 Hallazgos detectados")
            for nod in validos:
                cp3   = nod["conf_max"] * 100
                icono = "🔴" if nod["nivel"]=="alto" else "🟠" if nod["nivel"]=="medio" else "🟡"
                st.markdown(
                    f'<div class="{nod["css"]}">'
                    f'<div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;">'
                    f'<span style="color:{nod["color"]};font-weight:700;">{icono} Hallazgo {nod["id"]} — {nod["etiqueta"]}</span>'
                    f'<span style="color:{nod["color"]};font-weight:600;">Conf: {cp3:.1f}%</span>'
                    f'<span style="color:#8B949E;font-size:0.85rem;">{nod["tamano_estimado"]} — ~{nod["diametro_est_mm"]}mm</span>'
                    f'</div>'
                    f'<div style="color:#8B949E;font-size:0.82rem;margin-top:6px;display:flex;gap:20px;flex-wrap:wrap;">'
                    f'<span>📍 Rango: <b style="color:#E6EDF3;">slice_{nod["slice_inicio"]:03d} → slice_{nod["slice_fin"]:03d}</b></span>'
                    f'<span>🎯 Central: <b style="color:#E6EDF3;">slice_{nod["slice_central"]:03d}</b></span>'
                    f'<span>📊 {nod["n_slices"]} cortes / {nod["n_slices_rango"]} en rango</span>'
                    f'</div>'
                    f'<div style="color:#8B949E;font-size:0.8rem;margin-top:4px;">{nod["descripcion"]}</div>'
                    f'</div>', unsafe_allow_html=True)
                rc = next((r for r in res if r["nombre"]==nod["nombre_central"]), None)
                if rc:
                    ci, ci2 = st.columns([1,2])
                    with ci: st.image(rc["img_result"], caption=f"Corte central — {nod['nombre_central']}", use_container_width=True)
                    with ci2:
                        st.markdown(f'<div style="padding:0.5rem;color:#8B949E;font-size:0.82rem;">'
                                    f'<p><b style="color:#E6EDF3;">Clasificacion:</b> {nod["etiqueta"]}</p>'
                                    f'<p><b style="color:#E6EDF3;">Confianza promedio:</b> {nod["conf_prom"]*100:.1f}%</p>'
                                    f'<p style="color:#484F58;font-size:0.78rem;">Estimacion aproximada — confirmar con radiologo.</p></div>',
                                    unsafe_allow_html=True)

        if descartados:
            st.markdown("---")
            with st.expander(f"⚪ {len(descartados)} deteccion(es) descartada(s)", expanded=False):
                st.markdown('<div style="color:#9CA3AF;font-size:0.82rem;margin-bottom:0.8rem;">No superaron los criterios de validacion clinica.</div>', unsafe_allow_html=True)
                for nod in descartados:
                    st.markdown(f'<div class="fp-card"><span style="color:#9CA3AF;font-weight:600;">slice_{nod["slice_inicio"]:03d}</span>'
                                f'<span style="color:#6B7280;font-size:0.8rem;margin-left:12px;">{nod.get("razon_descarte","")}</span></div>',
                                unsafe_allow_html=True)

        st.markdown("---"); st.markdown("### 💾 Descargar reportes")
        dc1, dc2 = st.columns(2)
        with dc1:
            eb = generar_excel(res, mpac, validos, descartados)
            st.download_button("📊 Descargar Excel completo", data=eb,
                file_name=f"analisis_pulmonar_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with dc2:
            tb = generar_txt(res, mpac, validos, descartados, modo_ext)
            # Extraer nombre del paciente del primer archivo analizado
            nombre_paciente = "paciente_desconocido"
            if res:
                primer_archivo = res[0]["nombre"]
                m = re.match(r"(paciente_\d+)_", primer_archivo)
                if m:
                    nombre_paciente = m.group(1)

            st.download_button("📄 Descargar Reporte TXT", data=tb.encode("utf-8"),
                file_name=f"{nombre_paciente}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.txt",
                mime="text/plain", use_container_width=True)

        st.markdown("---"); st.markdown("#### 👤 Resumen del paciente")
        for pid, m in mpac.items():
            icono_pac = "🔴" if nivel_max==3 else "🟠" if nivel_max==2 else "🟡" if nivel_max==1 else "✅"
            etiq_pac  = ("Nodulo probable" if nivel_max==3 else "Sospecha" if nivel_max==2
                         else "Hallazgo incidental" if nivel_max==1 else "Sin hallazgos")
            with st.expander(f"📁 {pid} — {icono_pac} {etiq_pac}", expanded=True):
                pc = st.columns(6)
                vc = "#F87171" if nivel_max==3 else "#FB923C" if nivel_max==2 else "#FCD34D" if nivel_max==1 else "#4ADE80"
                with pc[0]: st.markdown(f'<div class="pac-metric-card"><p class="pac-metric-value" style="color:{vc};">{len(validos)}</p><p class="pac-metric-label">Hallazgos</p></div>',unsafe_allow_html=True)
                with pc[1]: st.markdown(f'<div class="pac-metric-card"><p class="pac-metric-value" style="color:#F87171;">{n_alto}</p><p class="pac-metric-label">Nodulo prob.</p></div>',unsafe_allow_html=True)
                with pc[2]: st.markdown(f'<div class="pac-metric-card"><p class="pac-metric-value" style="color:#FB923C;">{n_medio}</p><p class="pac-metric-label">Sospecha</p></div>',unsafe_allow_html=True)
                with pc[3]: st.markdown(f'<div class="pac-metric-card"><p class="pac-metric-value" style="color:#FCD34D;">{n_bajo}</p><p class="pac-metric-label">Hallazgo incid.</p></div>',unsafe_allow_html=True)
                with pc[4]: st.markdown(f'<div class="pac-metric-card"><p class="pac-metric-value" style="color:#3B82F6;">{cmg*100:.1f}%</p><p class="pac-metric-label">Conf max</p></div>',unsafe_allow_html=True)
                with pc[5]: st.markdown(f'<div class="pac-metric-card"><p class="pac-metric-value" style="color:#A78BFA;">{m["pct_afectado"]:.1f}%</p><p class="pac-metric-label">% cortes pos</p></div>',unsafe_allow_html=True)
                st.progress(m["pct_afectado"]/100)

        if sp:
            st.markdown("---"); st.markdown("### 🔬 Cortes con deteccion")
            for i in range(0, len(sp), 3):
                gr = sp[i:i+3]; cols = st.columns(3)
                for j, r in enumerate(gr):
                    cp3 = r["confianza"]*100
                    lab = "🔴" if cp3>=60 else "🟠" if cp3>=45 else "🟡"
                    with cols[j]:
                        st.image(r["img_result"], caption=f"{lab} {r['nombre']} — {cp3:.1f}%", use_container_width=True)
                        st.progress(r["confianza"])

        st.markdown("---")
        with st.expander("📊 Tabla completa de cortes", expanded=False):
            cn2, cr2, cc2 = st.columns([3,3,2])
            cn2.markdown("**Corte**"); cr2.markdown("**Clasificacion**"); cc2.markdown("**Confianza**")
            st.markdown("<hr style='margin:4px 0;border-color:#21262D'>", unsafe_allow_html=True)
            for r in res:
                cn2, cr2, cc2 = st.columns([3,3,2])
                cn2.markdown(f"<span style='color:#E6EDF3;font-size:0.85rem;'>{r['nombre']}</span>", unsafe_allow_html=True)
                cp3 = r["confianza"]*100
                if r["hay_nodulo"]:
                    if cp3>=60: cr2.markdown("<span class='nodule-badge'>ESTRUCTURA SOSPECHOSA</span>", unsafe_allow_html=True)
                    elif cp3>=45: cr2.markdown("<span class='sospecha-badge'>HALLAZGO INCIDENTAL</span>", unsafe_allow_html=True)
                    else: cr2.markdown("<span class='hallazgo-badge'>POSIBLE ARTEFACTO</span>", unsafe_allow_html=True)
                    cc2.markdown(f"<span style='color:#F87171;font-weight:600;'>{cp3:.1f}%</span>", unsafe_allow_html=True)
                else:
                    cr2.markdown("<span class='clean-badge'>SIN HALLAZGOS</span>", unsafe_allow_html=True)
                    cc2.markdown(f"<span style='color:#4ADE80;font-size:0.85rem;'>{cp3:.1f}%</span>", unsafe_allow_html=True)

        if sl_:
            with st.expander(f"Ver {len(sl_)} cortes sin hallazgos", expanded=False):
                for i in range(0, len(sl_), 4):
                    gr = sl_[i:i+4]; cols = st.columns(4)
                    for j, r in enumerate(gr):
                        with cols[j]: st.image(r["img_orig"], caption=f"✅ {r['nombre']}", use_container_width=True)

        st.markdown("---")
        st.markdown("""<div style="background:rgba(251,191,36,0.08);border:1px solid rgba(251,191,36,0.25);
        border-radius:10px;padding:1rem 1.2rem;color:#FDE68A;font-size:0.88rem;">
        ⚠️ <b>Aviso medico importante:</b> Este sistema es una herramienta computacional de apoyo al diagnostico,
        desarrollada como Proyecto de Grado academico. <b>NO constituye un diagnostico medico</b>.
        Todos los hallazgos deben ser confirmados por un radiologo o medico especialista certificado.
        </div>""", unsafe_allow_html=True)

else:
    st.markdown("""<div style="text-align:center;padding:3rem 1rem;color:#8B949E;">
    <div style="font-size:5rem;margin-bottom:1rem;">🫁</div>
    <p style="font-size:1.2rem;color:#E6EDF3;font-weight:600;">Sube los archivos del paciente para comenzar</p>
    <p style="font-size:0.9rem;max-width:560px;margin:0 auto;">
    Acepta <b style="color:#7FBBDB;">PNG/JPG</b>, <b style="color:#A5B4FC;">NPY</b>
    y <b style="color:#5EEAD4;">DICOM (.dcm o .zip)</b> — compatible con LUNA16, LIDC-IDRI y otros datasets</p>
    <div style="display:flex;justify-content:center;gap:1.2rem;flex-wrap:wrap;margin-top:2rem;">
    <div style="background:#161B22;border:1px solid #21262D;padding:1rem 1.5rem;border-radius:12px;min-width:120px;">
    <p style="color:#7FBBDB;font-size:1.1rem;font-weight:700;margin:0;">PNG/JPG</p>
    <p style="color:#8B949E;font-size:0.75rem;margin:0;">Slices exportados</p></div>
    <div style="background:#161B22;border:1px solid #21262D;padding:1rem 1.5rem;border-radius:12px;min-width:120px;">
    <p style="color:#A5B4FC;font-size:1.1rem;font-weight:700;margin:0;">NPY</p>
    <p style="color:#8B949E;font-size:0.75rem;margin:0;">Volumen completo</p></div>
    <div style="background:#161B22;border:1px solid #21262D;padding:1rem 1.5rem;border-radius:12px;min-width:120px;">
    <p style="color:#5EEAD4;font-size:1.1rem;font-weight:700;margin:0;">DICOM ZIP</p>
    <p style="color:#8B949E;font-size:0.75rem;margin:0;">Cualquier dataset</p></div>
    <div style="background:#EAB308;border:1px solid #CA8A04;padding:1rem 1.5rem;border-radius:12px;min-width:120px;">
    <p style="color:#1C1917;font-size:1rem;font-weight:700;margin:0;">⚡ Modo externo</p>
    <p style="color:#1C1917;font-size:0.75rem;margin:0;">Para LIDC-IDRI y datos clinicos</p></div>
    </div></div>""", unsafe_allow_html=True)

st.markdown("""<div class="footer-note">Sistema de Analisis Pulmonar — Proyecto de Grado 2026<br>
LUNA16 · YOLOv8 Triple Ensemble · NMS 3D · PNG · NPY · DICOM · CPU<br>
<span style="color:#3B82F6;">Herramienta de apoyo al diagnostico. No reemplaza la evaluacion medica profesional.</span>
</div>""", unsafe_allow_html=True)
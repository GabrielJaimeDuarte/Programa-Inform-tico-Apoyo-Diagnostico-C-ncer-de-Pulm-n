import os
import json
import random
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# RUTAS
# ==========================================
DIR_NPY_CON     = r"D:\IMAGENES PROYECTO DE GRADO\DATASET_NPY\con_nodulos"
DIR_NPY_SIN     = r"D:\IMAGENES PROYECTO DE GRADO\DATASET_NPY\sin_nodulos"
CSV_PACIENTES   = r"D:\IMAGENES PROYECTO DE GRADO\DATASET_NPY\nodulos_por_paciente.csv"
CSV_ANNOTATIONS = r"D:\IMAGENES PROYECTO DE GRADO\annotations.csv"
SALIDA_EXCEL    = r"D:\IMAGENES PROYECTO DE GRADO\nodulos_slices_exactos.xlsx"

# ==========================================
# ESTILOS
# ==========================================
h_fill  = PatternFill("solid", fgColor="0F3460")
h_font  = Font(bold=True, color="FFFFFF", size=10)
borde   = Border(
    left=Side(style="thin", color="2A3A4A"),
    right=Side(style="thin", color="2A3A4A"),
    top=Side(style="thin", color="2A3A4A"),
    bottom=Side(style="thin", color="2A3A4A"),
)
rojo_fill  = PatternFill("solid", fgColor="2A0808")
rojo_alt   = PatternFill("solid", fgColor="1A0505")
gris_fill  = PatternFill("solid", fgColor="161B22")
gris_alt   = PatternFill("solid", fgColor="0D1117")
verde_fill = PatternFill("solid", fgColor="0A1A0A")
verde_alt  = PatternFill("solid", fgColor="071207")


def leer_meta(nombre_paciente):
    for carpeta in [DIR_NPY_CON, DIR_NPY_SIN]:
        ruta = os.path.join(carpeta, f"{nombre_paciente}_meta.json")
        if os.path.exists(ruta):
            with open(ruta, "r", encoding="utf-8") as f:
                return json.load(f)
    return None


def obtener_slices_nodulo(nombre_paciente, coord_z):
    meta = leer_meta(nombre_paciente)
    if meta is None:
        return None, None, None, None

    shape_z   = meta.get("shape_ZYX", [0])[0]
    spacing_z = meta.get("spacing_XYZ_mm", [1.0, 1.0, 1.25])[2]

    if "annotations" in meta and "nodulos" in meta["annotations"]:
        mejor_nod  = None
        mejor_dist = 999999
        for nod in meta["annotations"]["nodulos"]:
            vz   = int(nod["voxel_Z"])
            dist = abs(vz * spacing_z - abs(coord_z))
            if dist < mejor_dist:
                mejor_dist = dist
                mejor_nod  = nod

        if mejor_nod:
            vz        = int(mejor_nod["voxel_Z"])
            z_menos1  = max(0, vz - 1)
            z_central = vz
            z_mas1    = min(shape_z - 1, vz + 1)
            return z_menos1, z_central, z_mas1, shape_z

    return None, None, None, shape_z


def generar_excel():
    print("Cargando datos...")
    pacs = pd.read_csv(CSV_PACIENTES)
    ann  = pd.read_csv(CSV_ANNOTATIONS)

    # Calcular split TRAIN/VAL con la misma semilla del preprocesador
    todos_pacientes = sorted(pacs['nombre_paciente'].tolist())
    random.seed(42)
    random.shuffle(todos_pacientes)
    split_idx = int(len(todos_pacientes) * 0.8)
    set_train = set(todos_pacientes[:split_idx])
    set_val   = set(todos_pacientes[split_idx:])
    pacs['split'] = pacs['nombre_paciente'].apply(
        lambda x: "VAL" if x in set_val else "TRAIN"
    )

    n_train = len(set_train)
    n_val   = len(set_val)
    print(f"  Split: {n_train} TRAIN | {n_val} VAL")

    merged = pacs.merge(ann, left_on='seriesuid_original',
                        right_on='seriesuid', how='left')

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════
    # HOJA 1 — Detalle nodulo por nodulo con slices exactos
    # ══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Nodulos con slices exactos"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:L1")
    c = ws1["A1"]
    c.value     = "Detalle de Nodulos con Slices Exactos - Dataset LUNA16"
    c.font      = Font(bold=True, color="FFFFFF", size=14)
    c.fill      = PatternFill("solid", fgColor="0F3460")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:L2")
    c2 = ws1["A2"]
    c2.value = (f"Total nodulos: {len(ann)}  |  "
                f"Pacientes TRAIN: {n_train}  |  "
                f"Pacientes VAL: {n_val}  |  "
                f"Split 80/20 seed=42")
    c2.font      = Font(color="7FBBDB", size=10)
    c2.fill      = PatternFill("solid", fgColor="0D1117")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 18

    enc = ["Paciente", "N nodulo", "Diametro (mm)", "Tamano",
           "Slice Z-1", "Slice Z central", "Slice Z+1",
           "Slices PNG generados", "coordZ (mm)",
           "Total slices CT", "Subset", "Split train/val"]
    for col, e in enumerate(enc, 1):
        c = ws1.cell(row=4, column=col, value=e)
        c.font      = h_font
        c.fill      = h_fill
        c.border    = borde
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[4].height = 22

    nods_con = merged[merged['tiene_nodulos'] == True].copy()
    nods_con = nods_con.sort_values('nombre_paciente').reset_index(drop=True)

    idx_nod          = {}
    fila             = 5
    total_procesados = 0

    print(f"Procesando {len(nods_con)} nodulos de {nods_con['nombre_paciente'].nunique()} pacientes...")

    for _, row_data in nods_con.iterrows():
        pac     = row_data['nombre_paciente']
        coord_z = row_data['coordZ']
        diam    = round(float(row_data['diameter_mm']), 2)

        if pac not in idx_nod:
            idx_nod[pac] = 0
        n_nod = idx_nod[pac]
        idx_nod[pac] += 1

        if diam < 6:    tam = "Pequeno (<6mm)"
        elif diam < 10: tam = "Mediano (6-10mm)"
        else:           tam = "Grande (>10mm)"

        color_diam = "F87171" if diam >= 10 else "FCD34D" if diam >= 6 else "E6EDF3"

        z_m1, z_c, z_p1, shape_z = obtener_slices_nodulo(pac, coord_z)

        if z_c is not None:
            slices_png   = (f"{pac}_nodulo_{n_nod}_Z{z_m1}.png  |  "
                           f"{pac}_nodulo_{n_nod}_Z{z_c}.png  |  "
                           f"{pac}_nodulo_{n_nod}_Z{z_p1}.png")
            slice_m1_val = z_m1
            slice_c_val  = z_c
            slice_p1_val = z_p1
            color_slice  = "4ADE80"
        else:
            slices_png   = "Meta.json no encontrado"
            slice_m1_val = "N/A"
            slice_c_val  = "N/A"
            slice_p1_val = "N/A"
            color_slice  = "888888"

        split_pac   = "VAL" if pac in set_val else "TRAIN"
        color_split = "3B82F6" if split_pac == "VAL" else "F59E0B"
        fondo       = rojo_fill if fila % 2 == 0 else rojo_alt

        datos = [
            pac, f"Nodulo {n_nod}", diam, tam,
            slice_m1_val, slice_c_val, slice_p1_val,
            slices_png,
            round(float(coord_z), 2),
            shape_z if shape_z else "N/A",
            row_data['subset'],
            split_pac,
        ]

        for col, val in enumerate(datos, 1):
            c = ws1.cell(row=fila, column=col, value=val)
            c.fill   = fondo
            c.border = borde
            c.alignment = Alignment(
                horizontal="left" if col in [1, 4, 8] else "center",
                vertical="center",
                wrap_text=(col == 8)
            )
            if col == 3:
                c.font = Font(bold=(diam >= 10), color=color_diam, size=10)
            elif col in [5, 6, 7]:
                c.font = Font(bold=(col == 6), color=color_slice, size=10)
            elif col == 12:
                c.font = Font(bold=True, color=color_split, size=10)
            else:
                c.font = Font(color="E6EDF3", size=10)

        ws1.row_dimensions[fila].height = 28 if z_c is not None else 16
        fila += 1
        total_procesados += 1

        if total_procesados % 100 == 0:
            print(f"  Procesados {total_procesados}/{len(nods_con)}...")

    anchos1 = [18, 12, 14, 18, 12, 18, 12, 68, 14, 16, 10, 14]
    for col, ancho in enumerate(anchos1, 1):
        ws1.column_dimensions[get_column_letter(col)].width = ancho

    # ══════════════════════════════════════════════════════
    # HOJA 2 — Resumen por paciente
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Resumen por paciente")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:H1")
    c = ws2["A1"]
    c.value     = "Resumen de Nodulos por Paciente"
    c.font      = Font(bold=True, color="FFFFFF", size=14)
    c.fill      = PatternFill("solid", fgColor="0F3460")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    enc2 = ["Paciente", "Tiene nodulos", "N nodulos",
            "Slices totales CT", "Todos los slices con nodulo",
            "Diametros (mm)", "Subset", "Split train/val"]
    for col, e in enumerate(enc2, 1):
        c = ws2.cell(row=3, column=col, value=e)
        c.font      = h_font
        c.fill      = h_fill
        c.border    = borde
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[3].height = 22

    fila2 = 4
    print("Generando hoja de resumen por paciente...")

    for _, pac in pacs.iterrows():
        pac_id      = pac['nombre_paciente']
        split_pac   = "VAL" if pac_id in set_val else "TRAIN"
        color_split = "3B82F6" if split_pac == "VAL" else "F59E0B"

        if pac['tiene_nodulos']:
            fondo2 = verde_fill if fila2 % 2 == 0 else verde_alt
        else:
            fondo2 = gris_fill if fila2 % 2 == 0 else gris_alt

        todos_slices = []
        diametros    = []

        if pac['tiene_nodulos']:
            meta = leer_meta(pac_id)
            if meta and "annotations" in meta and "nodulos" in meta["annotations"]:
                for nod in meta["annotations"]["nodulos"]:
                    vz    = int(nod["voxel_Z"])
                    shape = meta.get("shape_ZYX", [999])[0]
                    z_m1  = max(0, vz - 1)
                    z_p1  = min(shape - 1, vz + 1)
                    todos_slices.extend([z_m1, vz, z_p1])
                    diametros.append(f"{round(nod['diameter_mm'], 2)}mm")

        slices_texto = " | ".join([str(s) for s in sorted(set(todos_slices))]) if todos_slices else "sin meta.json"
        diam_texto   = " | ".join(diametros) if diametros else "-"

        datos2 = [
            pac_id,
            "SI" if pac['tiene_nodulos'] else "NO",
            int(pac['n_nodulos_ann']),
            int(pac['shape_Z']),
            slices_texto,
            diam_texto,
            pac['subset'],
            split_pac,
        ]

        for col, val in enumerate(datos2, 1):
            c = ws2.cell(row=fila2, column=col, value=val)
            c.fill   = fondo2
            c.border = borde
            c.alignment = Alignment(
                horizontal="left" if col in [1, 5, 6] else "center",
                vertical="center", wrap_text=True
            )
            if col == 2:
                c.font = Font(bold=True,
                             color="22C55E" if pac['tiene_nodulos'] else "666666",
                             size=10)
            elif col == 3 and pac['tiene_nodulos']:
                c.font = Font(bold=True, color="F87171", size=10)
            elif col == 5 and pac['tiene_nodulos']:
                c.font = Font(color="4ADE80", size=9)
            elif col == 8:
                c.font = Font(bold=True, color=color_split, size=10)
            else:
                c.font = Font(color="E6EDF3", size=10)

        ws2.row_dimensions[fila2].height = 20 if todos_slices else 16
        fila2 += 1

    anchos2 = [18, 14, 12, 18, 65, 30, 10, 14]
    for col, ancho in enumerate(anchos2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = ancho

    # ══════════════════════════════════════════════════════
    # HOJA 3 — Estadisticas
    # ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Estadisticas")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:C1")
    c = ws3["A1"]
    c.value     = "Estadisticas del Dataset LUNA16"
    c.font      = Font(bold=True, color="FFFFFF", size=14)
    c.fill      = PatternFill("solid", fgColor="0F3460")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 32

    pacs_train = pacs[pacs['nombre_paciente'].isin(set_train)]
    pacs_val   = pacs[pacs['nombre_paciente'].isin(set_val)]
    nods_train = merged[merged['nombre_paciente'].isin(set_train) & merged['tiene_nodulos']]
    nods_val   = merged[merged['nombre_paciente'].isin(set_val)   & merged['tiene_nodulos']]

    stats = [
        ("GENERAL",                        None),
        ("Total pacientes",                 len(pacs)),
        ("Pacientes con nodulos",           int(pacs['tiene_nodulos'].sum())),
        ("Pacientes sin nodulos",           int((~pacs['tiene_nodulos']).sum())),
        ("Total nodulos anotados",          len(ann)),
        ("",                                None),
        ("SPLIT TRAIN / VAL",               None),
        ("Pacientes TRAIN (80%)",           n_train),
        ("Pacientes VAL (20%)",             n_val),
        ("Nodulos en TRAIN",                len(nods_train)),
        ("Nodulos en VAL",                  len(nods_val)),
        ("Pac. con nodulos en TRAIN",       int(pacs_train['tiene_nodulos'].sum())),
        ("Pac. con nodulos en VAL",         int(pacs_val['tiene_nodulos'].sum())),
        ("",                                None),
        ("NODULOS POR PACIENTE",            None),
        ("Con 1 nodulo",                    int((pacs['n_nodulos_ann']==1).sum())),
        ("Con 2 nodulos",                   int((pacs['n_nodulos_ann']==2).sum())),
        ("Con 3 nodulos",                   int((pacs['n_nodulos_ann']==3).sum())),
        ("Con 4-5 nodulos",                 int(((pacs['n_nodulos_ann']>=4)&(pacs['n_nodulos_ann']<=5)).sum())),
        ("Con 6+ nodulos",                  int((pacs['n_nodulos_ann']>=6).sum())),
        ("Maximo en 1 paciente",            int(pacs['n_nodulos_ann'].max())),
        ("",                                None),
        ("TAMANOS DE NODULOS",              None),
        ("Pequenos (<6mm)",                 int((ann['diameter_mm']<6).sum())),
        ("Medianos (6-10mm)",               int(((ann['diameter_mm']>=6)&(ann['diameter_mm']<10)).sum())),
        ("Grandes (>10mm)",                 int((ann['diameter_mm']>=10).sum())),
        ("Diametro promedio (mm)",          round(float(ann['diameter_mm'].mean()), 2)),
        ("Diametro minimo (mm)",            round(float(ann['diameter_mm'].min()), 2)),
        ("Diametro maximo (mm)",            round(float(ann['diameter_mm'].max()), 2)),
        ("",                                None),
        ("SLICES POR CT",                   None),
        ("Promedio slices por CT",          round(float(pacs['shape_Z'].mean()), 1)),
        ("Minimo slices",                   int(pacs['shape_Z'].min())),
        ("Maximo slices",                   int(pacs['shape_Z'].max())),
    ]

    for i, (label, val) in enumerate(stats, 3):
        if val is None and label:
            ws3.merge_cells(f"A{i}:C{i}")
            c = ws3[f"A{i}"]
            c.value     = label
            c.font      = Font(bold=True, color="7FBBDB", size=11)
            c.fill      = PatternFill("solid", fgColor="0F2A40")
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws3.row_dimensions[i].height = 22
        elif label == "":
            ws3.row_dimensions[i].height = 6
        else:
            f = gris_fill if i % 2 == 0 else gris_alt
            c1 = ws3.cell(row=i, column=1, value=label)
            c1.font      = Font(color="8B949E", size=10)
            c1.fill      = f
            c1.border    = borde
            c1.alignment = Alignment(horizontal="left", vertical="center")

            c2 = ws3.cell(row=i, column=2, value=val)
            if "TRAIN" in label:
                c2.font = Font(bold=True, color="F59E0B", size=11)
            elif "VAL" in label:
                c2.font = Font(bold=True, color="3B82F6", size=11)
            else:
                c2.font = Font(bold=True, color="E6EDF3", size=11)
            c2.fill      = f
            c2.border    = borde
            c2.alignment = Alignment(horizontal="center", vertical="center")
            ws3.row_dimensions[i].height = 18

    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 16

    wb.save(SALIDA_EXCEL)
    print(f"\nExcel guardado en:")
    print(f"  {SALIDA_EXCEL}")
    print(f"\n  Hoja 1: {total_procesados} nodulos con slices exactos y split")
    print(f"  Hoja 2: {len(pacs)} pacientes con todos sus slices y split")
    print(f"  Hoja 3: estadisticas incluyendo distribucion train/val")
    print(f"\n  Colores de split:")
    print(f"  AMARILLO = TRAIN ({n_train} pacientes)")
    print(f"  AZUL     = VAL   ({n_val} pacientes)")


if __name__ == "__main__":
    generar_excel()